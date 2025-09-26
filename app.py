import streamlit as st
import pandas as pd
import altair as alt
from openpyxl import load_workbook
from io import BytesIO
from pathlib import Path
from PIL import Image
import io
import base64
import traceback


def main():
    st.set_page_config(page_title='PVC Pipe Sizing Explorer', layout='wide')

    # Load pool water background image FIRST
    pool_water_css = ""
    try:
        pool_water_path = Path(__file__).parent / 'pool_water.jpg'
        if pool_water_path.exists():
            with open(pool_water_path, 'rb') as f:
                pool_water_bytes = f.read()
            pool_water_b64 = base64.b64encode(pool_water_bytes).decode('utf-8')
            pool_water_css = f"""
            /* Tranquil pool water background - loaded as base64 */
            .main .block-container::after {{
                content: '';
                position: fixed;
                top: 0;
                right: 0;
                width: 35%;
                height: 100vh;
                background-image: url('data:image/jpeg;base64,{pool_water_b64}');
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                opacity: 0.25;
                z-index: -1;
                pointer-events: none;
                filter: blur(0.5px) brightness(1.1) contrast(0.9);
                mix-blend-mode: soft-light;
            }}
            
            /* Add a gentle gradient overlay for even more tranquility */
            .main .block-container::before {{
                content: '';
                position: fixed;
                top: 0;
                right: 0;
                width: 35%;
                height: 100vh;
                background: linear-gradient(135deg, rgba(240, 248, 255, 0.3) 0%, rgba(176, 224, 255, 0.2) 100%);
                z-index: -2;
                pointer-events: none;
            }}
            """
        else:
            # Fallback if image not found - make it more visible for debugging
            pool_water_css = """
            /* Fallback gradient background - more visible */
            .main .block-container::after {
                content: '';
                position: fixed;
                top: 0;
                right: 0;
                width: 35%;
                height: 100vh;
                background: radial-gradient(ellipse at 50% 50%, rgba(11, 130, 191, 0.15) 0%, rgba(176, 224, 255, 0.08) 100%);
                z-index: -1;
                pointer-events: none;
            }
            /* Debug border to see if background area is working */
            .main .block-container::before {
                content: 'Pool Water Area';
                position: fixed;
                top: 10px;
                right: 10px;
                background: rgba(11, 130, 191, 0.8);
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                z-index: 1000;
                font-size: 12px;
                pointer-events: none;
            }
            """
    except Exception as e:
        # Fallback if any error
        pool_water_css = """
        /* Error fallback - simple gradient */
        .main .block-container::after {
            content: '';
            position: fixed;
            top: 0;
            right: 0;
            width: 35%;
            height: 100vh;
            background: linear-gradient(135deg, rgba(240, 248, 255, 0.3) 0%, rgba(176, 224, 255, 0.1) 100%);
            z-index: -1;
            pointer-events: none;
        }
        """

    # CSS for styling and background colors
    st.markdown(f"""
    <style>
        .stApp {{ background-color: #f0f8ff; color: #03263a; }}
        .css-18e3th9 {{background-color: #f0f8ff;}}
        .stButton>button {{background-color:#0b82bf; color: #ffffff}}
        .stDownloadButton>button {{background-color:#0b82bf; color: #ffffff}}
        .stSidebar {{background-color:#e6f3ff}}
        .stMarkdown div{{color: #03263a}}
        img.logo-right {{max-height:200px}}
        
        /* ALLOW POOL WATER TO SHOW THROUGH - Remove aggressive backgrounds */
        .main .block-container {{ background-color: transparent !important; }}
        .stContainer {{ background-color: transparent !important; }}
        div[data-testid="stVerticalBlock"] {{ background-color: transparent !important; }}
        div[data-testid="stHorizontalBlock"] {{ background-color: transparent !important; }}
        .element-container {{ background-color: rgba(240, 248, 255, 0.7) !important; }}
        section[data-testid="stSidebar"] {{ background-color: #e6f3ff !important; }}
        
        /* Keep sidebar blue but make main area transparent */
        div[role="main"] {{ background-color: transparent !important; }}
        section.main {{ background-color: transparent !important; }}
        .css-k1vhr4, .css-1d391kg {{ background-color: transparent !important; }}
        
        /* Hide annoying mobile sidebar button - NUCLEAR APPROACH */
        button[kind="header"] {{ display: none !important; }}
        .css-1kyxreq {{ display: none !important; }}
        [data-testid="collapsedControl"] {{ display: none !important; }}
        .css-1rs6os {{ display: none !important; }}
        .css-17lntkn {{ display: none !important; }}
        [aria-label="Open sidebar"] {{ display: none !important; }}
        
        /* ULTIMATE NUCLEAR - Hide ALL sidebar toggle buttons */
        button[data-testid="baseButton-headerNoPadding"] {{ display: none !important; }}
        .css-vk3wp9 {{ display: none !important; }}
        [data-testid="stSidebarNav"] button {{ display: none !important; }}
        .stSidebar > div:first-child button {{ display: none !important; }}
        button:has([data-testid="keyboard_double_arrow_right"]) {{ display: none !important; }}
        [data-testid="keyboard_double_arrow_right"] {{ display: none !important; }}
        button:has(.css-nahz7x) {{ display: none !important; }}
        .css-nahz7x {{ display: none !important; }}
        
        /* Hide the entire sidebar header area on mobile */
        @media (max-width: 768px) {{
            [data-testid="stSidebarNav"] {{ display: none !important; }}
            .css-1d391kg {{ display: none !important; }}
        }}
        
        /* Style our custom inputs button */
        button[data-testid="baseButton-secondary"]:has([aria-label*="Inputs"]) {{
            background: linear-gradient(135deg, #e8f4fd 0%, #b3d9ff 100%) !important;
            border: 1px solid #0b82bf !important;
            color: #0b82bf !important;
        }}
        
        /* Pool water background will be inserted dynamically */
        {pool_water_css}
        
        /* Ensure content stays readable over the background */
        .main .block-container {{
            position: relative;
            z-index: 1;
        }}
        
        /* Mobile Responsiveness */
        @media (max-width: 768px) {{
            .main .block-container {{ padding: 1rem !important; }}
            img.logo-right {{ max-height: 150px; }}
            table {{ font-size: 16px !important; }}
            .kbi-badge {{ font-size: 18px !important; }}
        }}
    </style>
    """, unsafe_allow_html=True)

    # Add pool water background directly as HTML - FULL BACKGROUND
    try:
        pool_water_path = Path(__file__).parent / 'pool_water.jpg'
        if pool_water_path.exists():
            with open(pool_water_path, 'rb') as f:
                pool_water_bytes = f.read()
            pool_water_b64 = base64.b64encode(pool_water_bytes).decode('utf-8')
            
            st.markdown(f'''
            <div id="pool-water-bg" style="
                position: fixed;
                top: 0;
                left: 0;
                width: 100%;
                height: 100vh;
                background-image: url('data:image/jpeg;base64,{pool_water_b64}');
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                opacity: 0.4;
                z-index: -1;
                pointer-events: none;
            "></div>
            <style>
            /* Force pool water to stay visible after re-renders */
            #pool-water-bg {{
                position: fixed !important;
                z-index: -1 !important;
                opacity: 0.4 !important;
            }}
            /* Keep main areas transparent after re-render */
            .main .block-container {{
                background: transparent !important;
            }}
            .stApp > header {{
                background: transparent !important;
            }}
            </style>
            ''', unsafe_allow_html=True)
        else:
            # Show a test gradient to confirm positioning works
            st.markdown('''
            <div style="
                position: fixed;
                top: 0;
                left: 0;
                width: 100%;
                height: 100vh;
                background: linear-gradient(45deg, rgba(255, 0, 0, 0.3) 0%, rgba(0, 255, 0, 0.3) 50%, rgba(0, 0, 255, 0.3) 100%);
                z-index: -50;
                pointer-events: none;
            "></div>
            <div style="
                position: fixed;
                top: 50%;
                left: 50%;
                transform: translate(-50%, -50%);
                background: rgba(255, 255, 255, 0.9);
                color: black;
                padding: 20px;
                border-radius: 10px;
                z-index: 1000;
                font-size: 16px;
                text-align: center;
            ">Pool water image not found - showing test gradient</div>
            ''', unsafe_allow_html=True)
    except Exception as e:
        # Show error in center of screen
        st.markdown(f'''
        <div style="
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            background: rgba(255, 0, 0, 0.9);
            color: white;
            padding: 20px;
            border-radius: 10px;
            z-index: 1000;
            font-size: 16px;
            text-align: center;
        ">Error loading pool water: {str(e)}</div>
        ''', unsafe_allow_html=True)

    # Custom font loading
    @st.cache_data
    def load_custom_font():
        """Load custom font from file"""
        import pathlib
        
        app_dir = pathlib.Path(__file__).parent.absolute()
        font_path = app_dir / 'Frank the Architect.TTF'
        
        if font_path.exists():
            with open(font_path, "rb") as f:
                font_data = f.read()
            return base64.b64encode(font_data).decode()
        return None

    # Header with optional logo (prefers workspace logo file if present)
    col_title, col_logo = st.columns([3, 1])
    with col_title:
        st.title('PVC Pipe Sizing')
    with col_logo:
        workspace_logo = None
        app_dir = Path(__file__).parent
        # Look for JBDG logo first
        jbdg_logo = app_dir / 'JBDG Logo-2.jpg'
        if jbdg_logo.exists():
            workspace_logo = str(jbdg_logo)
        else:
            # Fallback to other logo files
            for ext in ('png', 'jpg', 'jpeg'):
                logo_path = app_dir / f'logo.{ext}'
                if logo_path.exists():
                    workspace_logo = str(logo_path)
                    break
        if workspace_logo:
            st.image(workspace_logo, width=200)

    # Load frankthearchitect font from the app directory
    embedded_font_css = None
    try:
        frank_path = Path(__file__).parent / 'Frank the Architect.TTF'
        if frank_path.exists():
            with open(frank_path, 'rb') as f:
                font_bytes = f.read()
            font_b64 = base64.b64encode(font_bytes).decode('utf-8')
            embedded_font_css = f"@font-face {{ font-family: 'FrankTheArchitect'; src: url('data:font/truetype;base64,{font_b64}') format('truetype'); }}"
    except Exception:
        pass

    # Use default logo color
    logo_color = '#0b82bf'

    # Mobile-friendly Inputs button (replaces the annoying sidebar button)
    st.markdown('&nbsp;', unsafe_allow_html=True)
    if st.button('📱 Inputs', key='mobile_inputs'):
        st.info('''
        **Quick Inputs:**
        
        Use the sidebar (desktop) or expand below (mobile) to set:
        - Flow Rate (gpm)
        - Line Type (Suction/Discharge)
        - Advanced Thresholds (optional)
        ''')
        
        # Show key inputs inline for mobile
        col1, col2 = st.columns(2)
        with col1:
            mobile_flow = st.number_input('Flow Rate (gpm)', value=600.0, key='mobile_flow')
        with col2:
            mobile_line = st.selectbox('Line Type', ['Suction', 'Discharge'], key='mobile_line')

    # Inject FrankTheArchitect font CSS for all text
    font_css_block = embedded_font_css or "@font-face { font-family: 'FrankTheArchitect'; src: local('FrankTheArchitect'), local('Architect'); }"
    st.markdown(f"""
    <style>
    {font_css_block}
    html, body, [class*="css"], .stApp, .stMarkdown, h1, h2, h3, h4, h5, h6, p, div, span {{ font-family: FrankTheArchitect, 'Architect', monospace !important; }}
    .kpi-badge {{
        display:inline-block; padding:8px 14px; border-radius:8px; color: white; font-weight:700; font-size:20px; font-family: FrankTheArchitect, monospace;
    }}
    .logo-accent {{ background: linear-gradient(90deg, {logo_color}, #ffffff22); padding:6px; border-radius:8px }}
    /* ULTIMATE SOLUTION: Completely replace expandable sections with custom styling */
    .stExpander {{ 
        background: white !important; 
        border: 3px solid #0b82bf !important; 
        border-radius: 10px !important; 
        margin: 10px 0 !important;
        overflow: hidden !important;
        position: relative !important;
    }}
    .stExpander details {{ 
        background: white !important; 
        border: none !important;
        margin: 0 !important;
        padding: 0 !important;
    }}
    .stExpander details summary {{ 
        background: #f0f8ff !important; 
        padding: 15px !important; 
        font-weight: bold !important; 
        cursor: pointer !important;
        border: none !important;
        position: relative !important;
        z-index: 1000 !important;
    }}
    .stExpander details[open] summary {{ 
        background: #e6f3ff !important; 
        border-bottom: 2px solid #0b82bf !important;
    }}
    /* NUCLEAR APPROACH: Hide everything that could contain keyboard text */
    .stExpander *::before, .stExpander *::after {{ 
        display: none !important; 
        visibility: hidden !important; 
        content: "" !important; 
        font-size: 0 !important;
        opacity: 0 !important;
    }}
    .stExpander summary::marker {{ display: none !important; }}
    .stExpander summary::-webkit-details-marker {{ display: none !important; }}
    /* Hide any icons or symbols */
    .stExpander .material-icons, .stExpander .material-symbols-outlined {{ 
        display: none !important; 
        visibility: hidden !important;
        font-size: 0 !important;
    }}
    </style>
    """, unsafe_allow_html=True)

    

    # Load workbook from the app directory
    excel_path = Path(__file__).parent / 'pipe_sizing.xlsx'
    try:
        with open(excel_path, 'rb') as f:
            wb_bytes = f.read()
    except FileNotFoundError:
        st.error(f'`pipe_sizing.xlsx` not found at {excel_path}. Please ensure the file exists in the same directory as this app.')
        return

    # Load workbook (formulas and values)
    wb = load_workbook(filename=BytesIO(wb_bytes), data_only=False)
    wb_vals = load_workbook(filename=BytesIO(wb_bytes), data_only=True)

    # Use the first sheet automatically (no display needed)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]
    ws_vals = wb_vals[sheet]


    def find_label_row(label_text_prefix):
        """Find a row index where column A starts with label_text_prefix (case-insensitive)."""
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().lower().startswith(label_text_prefix.lower()):
                return r
        return None


    # Read inputs by label (robust to small layout shifts)
    flow_row = find_label_row('Flow Rate')
    line_type_row = find_label_row('Line Type')
    nu_row = find_label_row('Water')

    # fallback row numbers if detection fails (based on current workbook)
    if flow_row is None:
        flow_row = 5
    if line_type_row is None:
        line_type_row = 6
    if nu_row is None:
        nu_row = 8


    def get_cell_value(r, c):
        return ws_vals.cell(row=r, column=c).value

    flow_gpm = get_cell_value(flow_row, 2) or 100.0
    line_type = get_cell_value(line_type_row, 2) or 'Suction'

    # Find constants by scanning the sheet for known labels in column D/E
    constants = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=4).value  # column D often contains constant labels
        val = ws_vals.cell(row=r, column=5).value  # column E often contains the numeric value
        if isinstance(label, str):
            constants[label.strip()] = val

    gal_to_ft3 = constants.get('gal_to_ft^3', 0.133681)
    sec_per_min = constants.get('sec_per_min', 60.0)
    pi = constants.get('π', 3.141592653589793)

    # kinematic viscosity is often in column B of its label row
    nu = get_cell_value(nu_row, 2) or 1.1e-05

    st.sidebar.subheader('Inputs')
    flow_gpm = st.sidebar.number_input('Flow Rate (gpm)', value=float(flow_gpm))

    # Restrict line type to only Suction or Discharge
    line_type = st.sidebar.selectbox('Line Type', options=['Suction', 'Discharge'], index=0 if (str(line_type).lower().startswith('suction')) else 1, help='Choose Suction or Discharge')
    st.sidebar.write('Water ν (ft²/s):', nu)

    # Threshold controls (using buttons instead of expanders)
    st.sidebar.markdown('&nbsp;', unsafe_allow_html=True)
    if st.sidebar.button('⚙️ Advanced Thresholds', key='thresholds_btn'):
        st.session_state.show_thresholds = not st.session_state.get('show_thresholds', False)
    
    if st.session_state.get('show_thresholds', False):
        if line_type == 'Suction':
            default_design = 4.5
            default_line = 6.0
        else:
            default_design = 6.0
            default_line = 8.0
        design_limit = st.sidebar.number_input('Design limit (ft/s)', value=float(default_design), help='Velocity considered fully acceptable')
        line_limit = st.sidebar.number_input('Line limit (ft/s)', value=float(default_line), help='Upper limit for this line type; velocities above this are unacceptable')
        st.sidebar.caption('Status rule: Acceptable if ≤ Design limit; Above design limit if ≤ Line limit; Unacceptable otherwise')
    else:
        # Use defaults when not shown
        if line_type == 'Suction':
            design_limit = 4.5
            line_limit = 6.0
        else:
            design_limit = 6.0
            line_limit = 8.0
    
    st.sidebar.container()
    st.sidebar.markdown('---')
    
    st.sidebar.markdown('---')  # Add separator


    # Debug toggle
    st.sidebar.container()
    st.sidebar.markdown('&nbsp;', unsafe_allow_html=True)  # Add space  
    debug = st.sidebar.checkbox('Show debug info', value=False)
    st.sidebar.container()

    # Locate the pipe size table by finding the header row that starts with 'Nominal Size'
    table_header_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower().startswith('nominal size'):
            table_header_row = r
            break

    if table_header_row is None:
        st.error('Could not find the pipe size table header in the sheet. Ensure the header starts with "Nominal Size" in column A.')
        return

    # Read table rows until first blank nominal size
    rows = []
    r = table_header_row + 1
    while r <= ws.max_row:
        nominal = ws_vals.cell(row=r, column=1).value
        if nominal is None:
            break
        inner_d = ws_vals.cell(row=r, column=2).value
        rows.append({'nominal_in': nominal, 'inner_in': inner_d, 'row': r})
        r += 1

    if not rows:
        st.error('No pipe sizes found under the table header.')
        return

    # Compute derived values
    data = []
    for item in rows:
        D_in = float(item['inner_in'])
        nominal_size = float(item['nominal_in'])
        
        # FORCE SKIP 1 INCH PIPE - only process if >= 1.5 inches
        if nominal_size < 1.5:
            continue
            
        area_ft2 = (pi * (D_in / 12) ** 2) / 4
        velocity = (flow_gpm * gal_to_ft3 / sec_per_min) / area_ft2
        # status using thresholds
        status_suction = 'Acceptable' if velocity <= design_limit else ('Above design limit' if velocity <= line_limit else 'Unacceptable')
        status_return = 'Acceptable' if velocity <= design_limit else ('Above design limit' if velocity <= line_limit else 'Unacceptable')
        # Reynolds number: Re = V * D(ft) / nu
        Re = velocity * (D_in / 12) / nu
        data.append({'Nominal (in)': nominal_size,
                     'Inner D (in)': D_in,
                     'Area (ft^2)': area_ft2,
                     'Velocity (ft/s)': velocity,
                     'Suction Status': status_suction,
                     'Return Status': status_return,
                     'Reynolds': Re,
                     'row': item['row']})

    df = pd.DataFrame(data)

    # Show debug info if toggled
    if debug:
        st.sidebar.write('Detected rows', rows[:6])
        st.sidebar.write('Constants', constants)

    # Choose which status column to display based on line type
    status_col = 'Suction Status' if line_type == 'Suction' else 'Return Status'
    df['Status'] = df[status_col]

    # Determine recommended size using the chosen line_limit
    active_limit = float(line_limit)
    recommended_df = df[df['Velocity (ft/s)'] <= active_limit]
    recommended_size = None
    recommended_velocity = None
    if not recommended_df.empty:
        # Get the first (smallest) acceptable size
        first_acceptable = recommended_df.iloc[0]
        recommended_size = first_acceptable['Nominal (in)']
        recommended_velocity = first_acceptable['Velocity (ft/s)']

    st.subheader('Computed pipe table')

    # Use AgGrid for nicer interactive table with conditional text coloring
    # Create ABSOLUTE MINIMALIST table - just pipe size and velocity!
    display_df = df[['Nominal (in)', 'Velocity (ft/s)']].copy()
    # Round velocity to 2 decimal places for cleaner display
    display_df['Velocity (ft/s)'] = display_df['Velocity (ft/s)'].round(2)
    
    # RESTART WITH NEW COLUMN NAMES TO BREAK CACHING
    display_df = display_df.rename(columns={
        'Nominal (in)': 'Nominal Pipe Size (in)',
        'Velocity (ft/s)': 'Velocity of water (ft/s)'
    })
    
    # FILTER OUT 12 INCH PIPE - we don't need it!
    display_df = display_df[display_df['Nominal Pipe Size (in)'] != 12.0]
    
    # SCREW AGGRID - LET'S MAKE OUR OWN BEAUTIFUL TABLE!
    
    # Build the table using st.write instead of st.markdown
    html_table = """
    <div style="display: flex; justify-content: center; margin: 20px 0;">
        <table style="border-collapse: collapse; width: 90%; max-width: 700px; font-family: 'FrankTheArchitect', monospace; font-size: 18px; background-color: #fbfcff; border-radius: 12px; overflow: hidden; box-shadow: 0 6px 12px rgba(11, 130, 191, 0.15); border: 1px solid #d6e9ff;">
            <thead>
                <tr style="background: linear-gradient(135deg, #0b82bf 0%, #1e90ff 100%); color: white;">
                    <th style="padding: 18px 20px; text-align: center; font-weight: bold; border: none; font-size: 16px;">Nominal Pipe Size (in)</th>
                    <th style="padding: 18px 20px; text-align: center; font-weight: bold; border: none; font-size: 16px;">Velocity of water (ft/s)</th>
                </tr>
            </thead>
            <tbody>
    """
    
    # Add each row with color coding
    for _, row in display_df.iterrows():
        velocity = row['Velocity of water (ft/s)']
        
        # Color code the velocity
        if velocity > line_limit:
            color = '#d62728'  # Red
        elif velocity > design_limit:
            color = '#ff7f0e'  # Orange
        else:
            color = '#2ca02c'  # Green
            
        html_table += f'<tr style="border-bottom: 1px solid #e6f3ff; background-color: {"#f8fbff" if len([x for x in display_df.iterrows()][:display_df.index.get_loc(row.name) + 1]) % 2 == 0 else "#fbfcff"};"><td style="padding: 14px 20px; text-align: center; font-weight: bold; border: none;">{row["Nominal Pipe Size (in)"]}</td><td style="padding: 14px 20px; text-align: center; font-weight: bold; color: {color}; border: none;">{velocity}</td></tr>'
    
    html_table += """
            </tbody>
        </table>
    </div>
    """
    
    # Try using st.write instead of st.markdown
    st.write(html_table, unsafe_allow_html=True)
    
    # Show recommended size badge with color (green/yellow/red) and include 'in'
    def badge_for_velocity(vel):
        if vel <= design_limit:
            return ('#2ca02c', 'Acceptable')
        if vel <= line_limit:
            return ('#ff7f0e', 'Above design limit')
        return ('#d62728', 'Unacceptable')

    # Choose a representative velocity for the badge
    try:
        vel_candidates = df.loc[df['Velocity (ft/s)'] <= active_limit, 'Velocity (ft/s)']
        vel_for_badge = float(vel_candidates.min()) if not vel_candidates.empty else float(df['Velocity (ft/s)'].max())
    except Exception:
        vel_for_badge = float(df['Velocity (ft/s)'].max()) if not df.empty else 0.0
    badge_color, badge_label = badge_for_velocity(vel_for_badge)
    
    # CENTERED RECOMMENDATION UNDER TABLE - NOW IN THE RIGHT PLACE!
    st.markdown(f"""
    <div style='width: 100%; text-align: center; margin: 25px auto; padding: 16px; 
                background: linear-gradient(135deg, #e8f4fd 0%, #f0f8ff 100%); 
                border: 2px solid #b3d9ff; border-radius: 12px; 
                max-width: 700px; box-shadow: 0 4px 8px rgba(11, 130, 191, 0.1);'>
        <div style='margin-bottom: 8px; font-size: 18px; font-weight: bold; color: #0b82bf;'>Recommended Pipe Size</div>
        {f'<div style="font-size: 24px; font-weight: bold; color: {badge_color}; margin: 8px 0;">{recommended_size} in</div>' if recommended_size else '<div style="font-size: 24px; font-weight: bold; color: #d62728; margin: 8px 0;">None Available</div>'}
        {f'<div style="font-size: 16px; color: {badge_color}; font-weight: bold;">Velocity: {recommended_velocity:.2f} ft/s</div>' if recommended_velocity else ''}
    </div>
    """, unsafe_allow_html=True)
    def badge_for_velocity(vel):
        if vel <= design_limit:
            return ('#2ca02c', 'Acceptable')
        if vel <= line_limit:
            return ('#ff7f0e', 'Above design limit')
        return ('#d62728', 'Unacceptable')

    # Choose a representative velocity for the badge: the smallest velocity that meets the active_limit,
    # or the overall max if none meet the limit. This is clearer than the previous one-liner.
    try:
        vel_candidates = df.loc[df['Velocity (ft/s)'] <= active_limit, 'Velocity (ft/s)']
        vel_for_badge = float(vel_candidates.min()) if not vel_candidates.empty else float(df['Velocity (ft/s)'].max())
    except Exception:
        vel_for_badge = float(df['Velocity (ft/s)'].max()) if not df.empty else 0.0
    # Status text centered under the table with background box - AFTER variables are defined!
    st.markdown(f"""
    <div style='width: 100%; text-align: center; margin: 25px auto; padding: 16px; 
                background: linear-gradient(135deg, #f0f8ff 0%, #e6f3ff 100%); 
                border: 2px solid #b3d9ff; border-radius: 12px; 
                max-width: 700px; font-size: 16px; box-shadow: 0 4px 8px rgba(11, 130, 191, 0.1);'>
        <strong style='color: #0b82bf;'>Status:</strong> {badge_label} &nbsp;•&nbsp; <em style='color: #0b82bf;'>Limit used:</em> {active_limit} ft/s
    </div>
    """, unsafe_allow_html=True)

    # Status explanation (using buttons instead of expanders)
    st.container()
    st.markdown('&nbsp;', unsafe_allow_html=True)
    if st.button('📋 Show Status Explanation', key='status_explanation'):
        st.info('''
        **Status Guide:**
        - **Acceptable** 🟢: Velocity ≤ Design limit (fully acceptable)
        - **Above design limit** 🟡: Design limit < Velocity ≤ Line limit (use with caution)
        - **Unacceptable** 🔴: Velocity > Line limit (do not use)
        ''')
    
    st.container()
    st.markdown('---')

    # Diagram / chart of velocities vs size (improved visuals)
    st.subheader('Velocity vs Pipe Size')
    # Build the chart robustly: convert and sort data, handle empty cases, and catch Altair errors
    if df.empty:
        st.info('No pipe sizes available to chart.')
    else:
        try:
            # Ensure numeric velocity and nominal ordering
            df['Velocity (ft/s)'] = pd.to_numeric(df['Velocity (ft/s)'], errors='coerce')
            # sanitize infinite values (avoid chained-assignment inplace warning)
            df['Velocity (ft/s)'] = df['Velocity (ft/s)'].replace([float('inf'), float('-inf')], pd.NA)

            # report NaNs and dtypes if debug
            if debug:
                st.sidebar.write('DataFrame dtypes:')
                st.sidebar.write(df.dtypes.astype(str).to_dict())
                st.sidebar.write('DataFrame head:')
                st.sidebar.write(df.head().to_dict(orient='records'))
                st.sidebar.write('Velocity NaN count:', int(df['Velocity (ft/s)'].isna().sum()))

            # keep nominal as string for category axis but sort by numeric value properly
            try:
                df['_nominal_sort'] = pd.to_numeric(df['Nominal (in)'], errors='coerce')
                # Sort dataframe by nominal size for proper chart ordering
                df = df.sort_values('_nominal_sort').reset_index(drop=True)
            except Exception:
                df['_nominal_sort'] = range(len(df))
            df['_nominal_label'] = df['Nominal (in)'].astype(str)

            # drop rows without numeric velocities before charting
            chart_df = df.dropna(subset=['Velocity (ft/s)']).copy()
            # FILTER OUT 12 INCH PIPE FROM CHART TOO!
            chart_df = chart_df[chart_df['Nominal (in)'] != 12.0]
            
            if chart_df.empty:
                st.info('No valid numeric velocities to chart after cleaning the data. See Chart debug below for computed values.')
                with st.expander('Chart debug — computed velocities', expanded=True):
                    # show computed velocities and source fields so user can see why chart is empty
                    try:
                        display_df = df[['Nominal (in)', 'Inner D (in)', 'Velocity (ft/s)']].copy()
                        display_df['Velocity (ft/s)'] = display_df['Velocity (ft/s)'].map(lambda v: f"{v:.6f}" if pd.notna(v) else 'NaN')
                        st.table(display_df)
                    except Exception:
                        st.write(df.head(20).to_dict(orient='records'))
                if debug:
                    st.sidebar.write('Cleaned df (first 20 rows):')
                    st.sidebar.write(df.head(20).to_dict(orient='records'))
            else:
                # Create color-coded bar chart based on status
                try:
                    # Sort chart data by nominal size
                    chart_df = chart_df.sort_values('_nominal_sort').reset_index(drop=True)
                    
                    # Add colors based on velocity status
                    chart_df['Color'] = chart_df['Velocity (ft/s)'].apply(
                        lambda v: '#2ca02c' if v <= design_limit else ('#ff7f0e' if v <= line_limit else '#d62728')
                    )
                    
                    # Create chart data with colors
                    chart_data = pd.DataFrame({
                        'Pipe Size': chart_df['_nominal_label'],
                        'Velocity': chart_df['Velocity (ft/s)'],
                        'Color': chart_df['Color'],
                        'Sort Order': chart_df['_nominal_sort']
                    })
                    
                    # Use Altair for color-coded bars with proper sorting
                    chart = alt.Chart(chart_data).mark_bar().encode(
                        x=alt.X('Pipe Size:N', title='Nominal Size (in)', sort=alt.EncodingSortField(field='Sort Order', op='mean')),
                        y=alt.Y('Velocity:Q', title='Velocity (ft/s)'),
                        color=alt.Color('Color:N', scale=None)
                    ).properties(height=360)
                    
                    st.altair_chart(chart, use_container_width=True)
                    
                    # Add limit line info below chart
                    st.caption(f'🟢 Green: ≤ {design_limit} ft/s (Acceptable)  |  🟡 Yellow: ≤ {line_limit} ft/s (Above design)  |  🔴 Red: > {line_limit} ft/s (Unacceptable)')
                    altair_ok = True
                except Exception:
                    # Fallback to simple bar chart
                    chart_data = chart_df.sort_values('_nominal_sort').set_index('_nominal_label')['Velocity (ft/s)']
                    st.bar_chart(chart_data)
                    st.caption(f'Horizontal reference: {active_limit} ft/s limit')
                    altair_ok = True                # Chart rendered successfully - no additional stats needed
        except Exception:
            st.error('Could not render chart. Enable debug to see details.')
            if debug:
                st.sidebar.error('Chart exception:')
                st.sidebar.text(traceback.format_exc())

    # Export button only shown when we have data
    # removed CSV download as requested

    st.caption('Notes: thresholds and constants were read from the workbook where possible. Adjust inputs in the sidebar to recalculate.')


if __name__ == '__main__':
    main()

